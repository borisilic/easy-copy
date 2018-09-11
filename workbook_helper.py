import string
import sys
import openpyxl


def get_file(filename):
    if filename:
        return open_workbook(filename)
    else:
        return open_workbook('default.xlsx')


def open_workbook(name):
    try:
        print('Opened file in location: ' + str(name))
        return openpyxl.load_workbook(name)
    except OSError as e:
        print('Ran into OSError: ' + str(e))
        print('Could not find file.')
        input('Press enter to exit program')
        sys.exit()


def set_headings(sheet):
    headings = []
    for cell in sheet[1]:
        headings.append(cell.value)
    return headings


def to_uppercase(index):
    return string.ascii_letters[index].upper()


def set_products(headings, row_number, sheet):
    products = []
    for row in row_number:
        line = dict()
        for heading in headings:
            cell_value = sheet[to_uppercase(headings.index(heading)) + str(row)].value
            line[heading] = cell_value
        products.append(line)
    display_line(0, products[0])
    return products


def get_previous_line(line_number, products):
    try:
        line_number = line_number - 1
        if line_number < 0:
            raise IndexError
        line_info = products[line_number]
        display_line(line_number, line_info)
        return line_number
    except IndexError:
        line_number = line_number + 1
        print('Reached end of list.')
        return line_number


def get_next_line(line_number, products):
    try:
        line_number = line_number + 1
        line_info = products[line_number]
        display_line(line_number, line_info)
        return line_number
    except IndexError:
        line_number = line_number - 1
        print('Reached end of list.')
        return line_number


def display_line(number, info):
    print()
    print('=' * 50)
    print('Current line: ' + str(number + 1))
    pretty(info, 1)
    print('=' * 50)


def pretty(d, indent=0):
    for key, value in d.items():
        print('\t' * indent + str(key))
        if isinstance(value, dict):
            pretty(value, indent+1)
        else:
            print('\t' * (indent+1) + str(value))


def set_columns_to_copy(args, headings):
    if args:
        columns = list(args.upper())
        if len(columns) == 1:
            try:
                return iterate_over_array(columns, headings)
            except IndexError:
                print('Some/all of the columns specified have no values in this workbook')
                input('Press enter to exit program')
                sys.exit()
        if len(columns) == 2:
            try:
                return iterate_over_array(columns, headings)
            except IndexError:
                print('Some/all of the columns specified have no values in this workbook')
                input('Press enter to exit program')
                sys.exit()
        else:
            try:
                return iterate_over_array(columns, headings)
            except IndexError:
                print('Some/all of the columns specified have no values in this workbook')
                input('Press enter to exit program')
                sys.exit()
    else:
        columns = ['A']
        return iterate_over_array(columns, headings)


def iterate_over_array(columns, headings):
    headings_to_copy = []
    for letter in columns:
        headings_to_copy.append(headings[string.ascii_uppercase.find(letter)])
    return headings_to_copy


def get_sheet(sheet, workbook):
    if sheet:
        sheets = workbook.sheetnames
        try:
            int(sheet)
            try:
                return workbook[sheets[int(sheet) - 1]]
            except IndexError:
                print('Sheet number ' + str(sheet) + ' cannot be found in this workbook')
                input('Press enter to exit program')
                sys.exit()
        except ValueError:
            try:
                return workbook[sheet]
            except KeyError:
                print('Cannot find sheet labeled: ' + sheet)
                input('Press enter to exit')
                sys.exit()
    else:
        sheets = workbook.sheetnames
        return workbook[sheets[0]]
